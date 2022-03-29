
import openpyxl

import numpy as np

wb=openpyxl.load_workbook(r'D:\Users\Administrator\Desktop\抗体数据存放.xlsx')
ws=wb.active
蓝耳=r'D:\Users\Administrator\Desktop\20210425蓝耳.TXT'

猪瘟=r'D:\Users\Administrator\Desktop\20210425猪瘟.TXT'

伪狂犬gB=r'D:\Users\Administrator\Desktop\20210425gB.TXT'

伪狂犬gE=r'D:\Users\Administrator\Desktop\20210425gE.TXT'

圆环=r'D:\Users\Administrator\Desktop\20210425圆环.TXT'

口蹄疫=r'D:\Users\Administrator\Desktop\20210425口蹄疫.TXT'

prrsva=open(蓝耳)
gBa=open(伪狂犬gB)
gEa=open(伪狂犬gE)
csfva=open(猪瘟)
pcv2a=open(圆环)
fmdva=open(口蹄疫)
# asfva=open(r'D:\Users\Administrator\Desktop\20210321非洲猪瘟.TXT')


p=np.array(prrsva.read().split()[22:])#蓝耳的写入
p.shape=8,13
prrsv1=p.flatten('F')[8:]
m1=3
for prrsv in prrsv1:
    ws.cell(column=3,row=m1,value=prrsv)
    m1=m1+1
    if m1>96:
        break
#
E=np.array(gEa.read().split()[22:])#ge的写入
E.shape=8,13
ge1=E.flatten('F')[8:]
m2=3
for gE in ge1:
    ws.cell(column=11,row=m2,value=gE)
    m2=m2+1
    if m2>96:
        break

B=np.array(gBa.read().split()[22:])#gB的写入
B.shape=8,13
gB1=B.flatten('F')[8:]
m3=3
for gB in gB1:
    ws.cell(column=8,row=m3,value=gB)
    m3=m3+1
    if m3>96:
        break

cs=np.array(csfva.read().split()[22:])#猪瘟的写入
cs.shape=8,13
csfv1=cs.flatten('F')[8:]
m4=3
for csfv in csfv1:
    ws.cell(column=14,row=m4,value=csfv)
    m4=m4+1
    if m4>96:
        break
圆环=np.array(pcv2a.read().split()[22:])#猪瘟的写入
圆环.shape=8,13
圆环=圆环.flatten('F')[8:]
m5=3
for 圆环 in 圆环:
    ws.cell(column=20,row=m5,value=圆环)
    m5=m5+1
    if m5>96:
        break
口蹄疫=np.array(fmdva.read().split()[22:])#猪瘟的写入
口蹄疫.shape=8,13
口蹄疫=口蹄疫.flatten('F')[8:]
m6=3
for 口蹄疫 in 口蹄疫:
    ws.cell(column=17,row=m6,value=口蹄疫)
    m6=m6+1
    if m6>96:
        break
# asf=np.array(asfva.read().split()[22:])#猪瘟的写入
# asf.shape=8,13
# asfv1=asf.flatten('F')[8:]
# m5=3
# for asfv in asfv1:
#     ws.cell(column=23,row=m5,value=asfv)
#     m5=m5+1
#     if m5>35:
#         break

wb.save(r'D:\Users\Administrator\Desktop\20210425抗体.xlsx')
#



