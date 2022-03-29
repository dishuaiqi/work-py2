import openpyxl

import numpy as np
import os
文件目录=r'F:\FC'#蓝耳2	猪瘟3	伪狂犬gB4	伪狂犬gE5	口蹄疫6	圆环7	非洲猪瘟8
第一板=input('请插入第一板名称：')
文件名1=第一板+'.TXT'
第二板=input('请输入第二板名称：')
文件名2=第二板+'.TXT'
第三板=input('请输入第三板名称：')
文件名3=第三板+'.TXT'
# 文件名3='20210522非洲猪瘟3.TXT'
列=int(input('蓝耳2	猪瘟3	伪狂犬gB4	伪狂犬gE5	口蹄疫6	圆环7	非洲猪瘟8,放在哪一列：'))
x=int(input('百测输入23，idexx输入22：'))
wb=openpyxl.load_workbook(r'D:\Users\Administrator\Desktop\抗体数据.xlsx')
ws=wb.active
数据1=os.path.join(文件目录,文件名1)


数据1=open(数据1)
数据1=np.array(数据1.read().split()[x:])#蓝耳的写入
# print(数据1)
数据1.shape=8,13
数据1=数据1.flatten('F')[8:]
m1=2
for 数据1 in 数据1:
    ws.cell(column=列,row=m1,value=数据1)
    m1=m1+1
    if m1>97:
        break

数据2=os.path.join(文件目录,文件名2)
数据2=open(数据2)
数据2=np.array(数据2.read().split()[x:])#蓝耳的写入
数据2.shape=8,13
数据2=数据2.flatten('F')[8:]
m2=98
for 数据2 in 数据2:
    ws.cell(column=列,row=m2,value=数据2)
    m2=m2+1
    if m2>193:
        break

数据3=os.path.join(文件目录,文件名3)
数据3=open(数据3)
数据3=np.array(数据3.read().split()[x:])#蓝耳的写入
数据3.shape=8,13
数据3=数据3.flatten('F')[8:]
m3=194
for 数据3 in 数据3:
    ws.cell(column=列,row=m3,value=数据3)
    m3=m3+1
    if m3>289:
        break




# 文件名3='20210505口蹄疫3.TXT'
# 数据3=os.path.join(文件目录,文件名3)
# 数据3=open(数据3)
# 数据3=np.array(数据3.read().split()[x:])#蓝耳的写入
# 数据3.shape=8,13
# 数据3=数据3.flatten('F')[8:]
# m3=194
# for 数据3 in 数据3:
#     ws.cell(column=列,row=m3,value=数据3)
#     m3=m3+1
#     if m3>289:
#         break
# 文件名4='20210505口蹄疫4.TXT'
# 数据4=os.path.join(文件目录,文件名4)
# 数据4=open(数据4)
# 数据4=np.array(数据4.read().split()[x:])#蓝耳的写入
# 数据4.shape=8,13
# 数据4=数据4.flatten('F')[8:]
# m4=290
# for 数据4 in 数据4:
#     ws.cell(column=列,row=m4,value=数据4)
#     m4=m4+1
#     if m4>386:
#         break

wb.save(r'D:\Users\Administrator\Desktop\抗体数据.xlsx')